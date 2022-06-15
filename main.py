from tkinter import filedialog
import cv2
import numpy as np
import operator
import glob
import openpyxl as pyxl
from datetime import datetime
import tkinter as tk
import tkinter.font

class CorrectionAuto:
    paquetfile = ""         ## valeur de stockage du lien du paquet de copie
    corrfile = ""           ## valeur de stockage du lien de correction
    nbquestion = 0          ##nombre de question dans le qcm
    tabresultat = [[0, 0]]  ## init tab de résultat final
    UE = False     ## notation différente si ue différente
    version= "Version 1.2"

    def __init__(self):
        self.paquetfile = ""
        self.corrfile = ""
        self.nbquestion = 0
        self.tabresultat = [[0, 0]]
        self.done = False

    def ChangerUE(self,BIN):
        self.UE = BIN

    def AffilierCorr(self,file ):       ## deff d'affiliation de correctionfile Appellée apres pression sur le bouton
            self.corrfile = file

    def AffilierDoss(self, file):       ## def d'affiliation de paquetfile Apellée apres pression dur le bouton
        self.paquetfile = file

    def AffilierNBQ(self, nombre):         ## same pour nbquestion
        self.nbquestion = nombre

    def Correction(self):                   ## def principale de correction du qcm
        if self.done==True:         ## test pour eviter de corriger deux fois le meme qcm
            pass
        else:


            ##################### PARTIE VARIABLES #####################

            largeurtabEtu = 1140
            hauteurtabEtu = 935
            NombreQuestion = self.nbquestion

            ImgCorr = cv2.imread(self.corrfile, 0)
            methode = cv2.ADAPTIVE_THRESH_GAUSSIAN_C  ## methode pour le seuillage

            v1 = 9  ##

            hauteurtab = 1444  ## Valeurs pour le tab de persepective
            largeurtab = 502

            ListImg = []        ##tab de stockage des images du paquet de copie

            file_list = glob.glob(self.paquetfile + '/*.*')     ## recherche et mise en tab des images dans le dossier donné

            for file in file_list:
                a = cv2.imread(file, 0)
                ListImg.append(a)

            widthGrille = 502  ## redimension de la grille uniquement
            heightGrille = 1444

            tabcentre = np.array([[33, 109, 180, 251, 322, 393, 464],## 1     ## tab des centres des cases du qcm avec premiere valeur est l'ordonnée
                                  [80, 109, 180, 251, 322, 393, 464],##2
                                  [129, 109, 180, 251, 322, 393, 464],#3
                                  [175, 109, 180, 251, 322, 393, 464],#4
                                  [223, 109, 180, 251, 322, 393, 464],#5
                                  [270, 109, 180, 251, 322, 393, 464],#6
                                  [317, 109, 180, 251, 322, 393, 464],#7
                                  [366, 109, 180, 251, 322, 393, 464],#8
                                  [411, 109, 180, 251, 322, 393, 464],#9
                                  [459, 109, 180, 251, 322, 393, 464],#10
                                  [507, 109, 180, 251, 322, 393, 464],#11
                                  [554, 109, 180, 251, 322, 393, 464],#12
                                  [602, 109, 180, 251, 322, 393, 464],#13
                                  [650, 109, 180, 251, 322, 393, 464],  #14
                                  [700, 109, 180, 251, 322, 393, 464],#15
                                  [744, 109, 180, 251, 322, 393, 464],#16
                                  [792, 109, 180, 251, 322, 393, 464],#17
                                  [840, 109, 180, 251, 322, 393, 464],#18
                                  [886, 109, 180, 251, 322, 393, 464],#19
                                  [935, 109, 180, 251, 322, 393, 464],#20
                                  [980, 109, 180, 251, 322, 393, 464],#21
                                  [1029, 109, 180, 251, 322, 393, 464],#22
                                  [1076, 109, 180, 251, 322, 393, 464],#23
                                  [1122, 109, 180, 251, 322, 393, 464],#24
                                  [1171, 109, 180, 251, 322, 393, 464],#25
                                  [1218, 109, 180, 251, 322, 393, 464],#26
                                  [1266, 109, 180, 251, 322, 393, 464],#27
                                  [1313, 109, 180, 251, 322, 393, 464],#28
                                  [1361, 109, 180, 251, 322, 393, 464],#29
                                  [1400, 109, 180, 251, 322, 393, 464]])#30

            tabcentreEtu = np.array([[54, 74, 216, 359, 502, 642, 781, 922, 1061],      ## pareil pour centre dans le tableau des numeros etud
                                     [147, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [243, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [332, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [421, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [507, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [600, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [684, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [776, 74, 216, 359, 502, 642, 781, 922, 1061],
                                     [861, 74, 216, 359, 502, 642, 781, 922, 1061]])

            ##################### FIN PARTIE VARIABLES #####################

            ##################### PARTIE CORRECTION #####################

            grayCorr = cv2.GaussianBlur(ImgCorr, (3, 3), 0)  ##ajout de flou

            threshCorr = cv2.adaptiveThreshold(grayCorr, 255, methode, cv2.THRESH_BINARY_INV, v1, 2)  # seuillage avec une valeur de 255

            contoursCorr, hierarchy = cv2.findContours(threshCorr, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)  ## methode de recheerche de contour

            maxAreaCorr = 0

            for c in contoursCorr:  ## boucle de creation d'une bordure
                area = cv2.contourArea(c)
                if area > 250000:
                    peri = cv2.arcLength(c, True)
                    polygoneCorr = cv2.approxPolyDP(c, 0.01 * peri, True)
                    if area > maxAreaCorr and len(polygoneCorr) == 4:
                        contour_grille = polygoneCorr
                        maxAreaCorr = area

            if contour_grille is not None:  ## dessin de bordure quand bordure trouvée
                cv2.drawContours(ImgCorr, [contour_grille], 0, (0, 255, 0), 2)
                pointsCorr = np.vstack(contour_grille).squeeze()
                pointsCorr = sorted(pointsCorr, key=operator.itemgetter(1))

                if pointsCorr[0][0] < pointsCorr[1][0]:  ## serie de test pour tjr avoir la grille dans le bon sens

                    if pointsCorr[3][0] < pointsCorr[2][0]:
                        pts1Corr = np.float32([pointsCorr[0], pointsCorr[1], pointsCorr[3], pointsCorr[2]])
                    else:
                        pts1Corr = np.float32([pointsCorr[0], pointsCorr[1], pointsCorr[2], pointsCorr[3]])
                else:
                    if pointsCorr[3][0] < pointsCorr[2][0]:
                        pts1Corr = np.float32([pointsCorr[1], pointsCorr[0], pointsCorr[3], pointsCorr[2]])
                    else:
                        pts1Corr = np.float32([pointsCorr[1], pointsCorr[0], pointsCorr[2], pointsCorr[3]])

            pts2Corr = np.float32([[0, 0], [largeurtab, 0], [0, hauteurtab], [largeurtab, hauteurtab]])  ## initialisation pour 63
            MCorr = cv2.getPerspectiveTransform(pts1Corr, pts2Corr)  ## creation de la matrice pour la grille
            grilleCorr = cv2.warpPerspective(ImgCorr, MCorr, (largeurtab, hauteurtab))  ## creation de l'image grille
            grilleCorr = cv2.resize(grilleCorr, (widthGrille, heightGrille), interpolation=cv2.INTER_AREA)  ## resize

            ## cv2.imwrite('grilleCorr.png', grilleCorr)  ##creation de l'image du qcm uniquement debug

            tabCorr = np.array([[0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0],
                                [0, 0, 0, 0, 0, 0]])

            for ligne in range(int(NombreQuestion)):  ## boucle de remplissage de tabRep avec qcm
                for colone in range(6):
                    if grilleCorr[tabcentre[ligne][0]][tabcentre[ligne][colone + 1]] <= 100:
                        tabCorr[ligne][colone] = 1;

            self.tabresultat[0] = ['correction', int(NombreQuestion)]       ## ne pas laisser le 0,0 d'init
            print(" la partie correction a été effectué")       ## debug

            ##################### FIN PARTIE CORRECTION #####################

            ##################### PARTIE QCM  #####################

            nbEtudiant = len(ListImg)
            for numero in range(nbEtudiant):
                Img = ListImg[numero]           ## recupérention d'image dns une var

                        ## Img = cv2.resize(Img, (LCorr, HCorr), interpolation=cv2.INTER_AREA)
                        ##enchainenement de traitement sur l'image pour pouvoir trouver des contours facilement lors de l'appel de la fonction
                        ##   gray = cv2.GaussianBlur(Img, (3, 3), 0)  ##ajout de flou
                thresh = cv2.adaptiveThreshold(Img, 200, methode, cv2.THRESH_BINARY_INV, v1,2)  # seuillage avec une valeur de 255

                ##################### PARTIE NUMERO ETUDIANT  #####################

                contoursEtu, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)  ## methode de recheerche de contour
                contour_Etu = None  ##
                maxAreaEtu = 0

                for c in contoursEtu:  ## boucle de creation d'une bordure
                    areaEtu = cv2.contourArea(c)
                    if areaEtu > 100000 and areaEtu < 130000:

                        periEtu = cv2.arcLength(c, True)
                        polygoneEtu = cv2.approxPolyDP(c, 0.01 * periEtu, True)
                        if areaEtu > maxAreaEtu and len(polygoneEtu) == 4:
                            contour_Etu = polygoneEtu
                            maxAreaEtu = areaEtu

                if contour_Etu is not None:  ## dessin de bordure quand bordure trouvée
                    cv2.drawContours(Img, [contour_Etu], 0, (0, 255, 0), 2)
                    pointsEtu = np.vstack(contour_Etu).squeeze()
                    pointsEtu = sorted(pointsEtu, key=operator.itemgetter(1))

                    if pointsEtu[0][0] < pointsEtu[1][0]:  ## serie de test pour tjr avoir la grille dans le bon sens

                        if pointsEtu[3][0] < pointsEtu[2][0]:
                            pts1Etu = np.float32([pointsEtu[0], pointsEtu[1], pointsEtu[3], pointsEtu[2]])
                        else:
                            pts1Etu = np.float32([pointsEtu[0], pointsEtu[1], pointsEtu[2], pointsEtu[3]])
                    else:
                        if pointsEtu[3][0] < pointsEtu[2][0]:
                            pts1Etu = np.float32([pointsEtu[1], pointsEtu[0], pointsEtu[3], pointsEtu[2]])
                        else:
                            pts1Etu = np.float32([pointsEtu[1], pointsEtu[0], pointsEtu[2], pointsEtu[3]])

                    pts2Etu = np.float32(
                        [[0, 0], [largeurtabEtu, 0], [0, hauteurtabEtu],[largeurtabEtu, hauteurtabEtu]])  ## initialisation pour 63
                    MEtu = cv2.getPerspectiveTransform(pts1Etu, pts2Etu)  ## creation de la matrice pour la grille
                    Etudiant = cv2.warpPerspective(Img, MEtu,(largeurtabEtu, hauteurtabEtu))  ## creation de l'image grille
                    cv2.imwrite('NumeroEtu.png', Etudiant)
                    tabNE = np.array([[0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0],
                                      [0, 0, 0, 0, 0, 0, 0, 0]])

                    for ligneEtu in range(10):  ## boucle de remplissage de tabRep avec numero etudiant
                        for coloneEtu in range(8):
                            if Etudiant[tabcentreEtu[ligneEtu][0]][tabcentreEtu[ligneEtu][coloneEtu + 1]] <= 100:
                                tabNE[ligneEtu][coloneEtu] = 1;

                NumeroEtudiant = ""
                for coloneNE in range(8):
                    for ligneNE in range(10):
                        if tabNE[ligneNE][coloneNE] != 0:
                            NumeroEtudiant = NumeroEtudiant + str(ligneNE)

                ##################### FIN PARTIE NUMERO ETUDIANT  #####################

                ##################### PARTIE CORRECTION QCM  #####################

                contours, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)  ## methode de recheerche de contour
                contour_grille = None  ##
                maxArea = 0

                for c in contours:  ## boucle de creation d'une bordure
                    area = cv2.contourArea(c)
                    if area > 25000:
                        peri = cv2.arcLength(c, True)
                        polygone = cv2.approxPolyDP(c, 0.01 * peri, True)
                        if area > maxArea and len(polygone) == 4:
                            contour_grille = polygone
                            maxArea = area

                if contour_grille is not None:  ## dessin de bordure quand bordure trouvée
                    cv2.drawContours(Img, [contour_grille], 0, (0, 255, 0), 2)
                    points = np.vstack(contour_grille).squeeze()
                    points = sorted(points, key=operator.itemgetter(1))

                    if points[0][0] < points[1][0]:  ## serie de test pour tjr avoir la grille dans le bon sens

                        if points[3][0] < points[2][0]:
                            pts1 = np.float32([points[0], points[1], points[3], points[2]])
                        else:
                            pts1 = np.float32([points[0], points[1], points[2], points[3]])
                    else:
                        if points[3][0] < points[2][0]:
                            pts1 = np.float32([points[1], points[0], points[3], points[2]])
                        else:
                            pts1 = np.float32([points[1], points[0], points[2], points[3]])

                    pts2 = np.float32([[0, 0], [largeurtab, 0], [0, hauteurtab],[largeurtab, hauteurtab]])  ## initialisation pour 63
                    M = cv2.getPerspectiveTransform(pts1, pts2)  ## creation de la matrice pour la grille
                    grille = cv2.warpPerspective(Img, M, (largeurtab, hauteurtab))  ## creation de l'image grille
                    cv2.imwrite('grille.png', grille)
                tabRep = np.array([[0, 0, 0, 0, 0, 0],          ## tableau de reponse 0 = non coché 1 = coché
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0],
                                   [0, 0, 0, 0, 0, 0]
                                   ])
                tabRepLigne = []
                for ligne in range ( int(NombreQuestion)):
                    tabRepLigne.append(0)
                for ligne in range(int(NombreQuestion)):  ## boucle de remplissage de tabRep avec qcm
                    for colone in range(6):
                        if grille[tabcentre[ligne][0]][tabcentre[ligne][colone + 1]] <= 100:
                            tabRep[ligne][colone] = 1;
                            tabRepLigne[ligne] = 1



                if (self.UE==True):            ## différente notation entre les ue
                    note = int(NombreQuestion)
                    for ligne in range(int(NombreQuestion)):
                        fauteligne = 0
                        if tabRepLigne[ligne]==0:
                            note = note -1
                        else:
                            for colone in range(6):
                                if tabCorr[ligne][colone] != tabRep[ligne][colone]:
                                    fauteligne = fauteligne + 1
                            if tabCorr[ligne][5] == tabRep[ligne][5]:       ## F = tout est vrai, donc si différent question fausse
                                if fauteligne == 1:
                                    note = note - 0.5
                                elif(fauteligne >=2):
                                    note = note -1           ## notation = une différence = 0
                            else:
                                note = note - 1
                    self.tabresultat.append([NumeroEtudiant, note])         ## ajout du numero etudiant sous forme str et de la note dans le tableau des résultats

                else:
                    note = int(NombreQuestion)
                    for ligne in range(int(NombreQuestion)):
                        fauteligne = 0
                        if tabRepLigne[ligne] == 0:
                            note = note -1
                        else:
                            for colone in range(6):
                                if tabCorr[ligne][colone] != tabRep[ligne][colone]:
                                    fauteligne = fauteligne + 1
                            if tabCorr[ligne][5] == tabRep[ligne][5]:
                                if fauteligne >= 3:
                                    note = note - 1
                                else:
                                    note = note - (fauteligne * 0.25)
                            else:
                                note = note - 1
                    self.tabresultat.append([NumeroEtudiant, note])

            self.done = True
            print("chaussure :) ") ##debug


    def toExcel(self,file):         ## def de mise en excel du tableau de résultat


        time  = str(datetime.now())         ## pour avoir le temps sur l'excel
        workbook = pyxl.Workbook()

        if (file == ""):        ## eviter de faire deux excel différents et tout recalculer
            pass
        else:
            sheet1 = workbook.active            ## création feuille dans le workbook
            sheet1.title = "Feuil1"
            sheet1.column_dimensions["A"].width = 20        ## taille tableau
            sheet1.column_dimensions["E"].width = 15
            ## corp d'excel
            sheet1['A1'].value = "NUMERO ETUDIANT"                          ## préremplissage des cases avec modèle donné par la fac
            sheet1['B1'].value = "NOTE"
            sheet1['C1'].value = "RANG"
            sheet1['E1'].value = "DATE :"
            sheet1['F1'].value = time
            sheet1['E2'].value = "Effectif"
            sheet1['F2'].value = "=COUNTA(Feuil1!$B$2:$B$412)"
            sheet1['E3'].value = "Moyenne"
            sheet1['F3'].value = "=AVERAGE(Feuil1!$B$2:$B$412)"
            sheet1['E4'].value = "Médiane"
            sheet1['F4'].value = "=MEDIAN(Feuil1!$B$2:$B$412)"
            sheet1['E5'].value = "Ecart type"
            sheet1['F5'].value = "=STDEV(Feuil1!$B$2:$B$412)"
            sheet1['E6'].value = "1er Quartile"
            sheet1['F6'].value = "=PERCENTILE(Feuil1!$B$2:$B$412,0.75)"
            sheet1['E7'].value = "1er centile"
            sheet1['F7'].value = "=PERCENTILE(Feuil1!$B$2:$B$412,0.9)"
            sheet1['E8'].value = "Max"
            sheet1['F8'].value = "=MAX(Feuil1!$B$2:$B$412)"
            sheet1['E9'].value = "MIN"
            sheet1['F9'].value = "=MIN(Feuil1!$B$2:$B$412)"
            ## Remplissage excel avec valeurs

            for indice in range(len(self.tabresultat)):                     ## remplissage des colones avec le tabrésultat du qcm
                if indice == 0:              ## ne pas mettre la premiere ligne qui correspond a la correction dans l'excel
                    pass
                else:
                    sheet1.cell(indice+1, 1).value = self.tabresultat[indice][0]               ## ajout numero etu
                    sheet1.cell(indice+1, 2).value = self.tabresultat[indice][1]               ## ajout note
                    test = str(indice +1)
                    formule = "=RANK(Feuil1!$B"+ test + ",Feuil1!$B$2:$B$412,0)"
                    sheet1.cell(indice+1, 3).value = formule  ## ajout rang

            workbook.save(file +'.xlsx')          ## sauvegarde du fichier a lendroit demandé à l'utilisateur


class Application(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.geometry('400x400')
        self.resizable(width=0, height=0)
        self.iconbitmap("Logo valid.ico")
        self.title("Qcm correction automatique")
        self.Qcm = CorrectionAuto()
        self.creer_widgets(self.Qcm)
        FON_Police = tkinter.font.Font(size=20)

    def creer_widgets(self, QCM : CorrectionAuto):
        self.canva1 = tk.Canvas(self, width=400, height=400)  ## canva principal qui fait toute la fenetre
        self.canva1.create_rectangle(20, 20, 380, 50, outline="black")  ## rectangle du haut, pour mettre titre
        self.canva1.create_text(200, 35, text="Correction Automatique QCM") ## Titre dans le rectangle
        self.canva1.place(x=0, y=0) ## placer le canva
        self.canva1.create_oval(360, 60, 390, 90, fill='red')   ## voyant check correction
        self.canva1.create_oval(360,105,390,135, fill = 'red')  ## voyant check Paquet de copie

        self.canva2 = tk.Canvas(self, width=30,height = 30)
        FON_Police = tkinter.font.Font(size=20)
        self.canva2.create_text(15,15, text="0",font=FON_Police) ## Titre dans le rectangle
        self.canva2.place(x=350,y=170)

        self.bouton_askCorr = tk.Button( text="Donner la Correction", command=lambda: QCM.AffilierCorr(self.AskCorrFile())).place(x=20, y=70)
        self.bouton_askDossCopies = tk.Button( text="Donner le paquet de copies", command=lambda: QCM.AffilierDoss(self.AskPackFile())).place(x=20, y=110)
        self.Entry1 = tk.Entry( width=20)
        self.Entry1.place(x=20, y=180)
        self.canva1.create_text(125, 160, text="Entrer nombre de question dans le qcm")
        self.bouton_NbQ = tk.Button( text="Soumettre nombre de questions", command=lambda: QCM.AffilierNBQ(self.AskNumber())).place(x=150, y=175)
        self.boutonCorrectionAuto = tk.Button( text="Lancer correction automatique", command=lambda: self.TestCorrection()).place(x=20, y=310)
        self.bouton_SaveExcel = tk.Button(self, text="Enregistrer le classement", command=lambda: self.AskExcelFile()).place(x=220, y=310)
        self.quitButon= tk.Button(self,text="Quitter", command=lambda :self.destroy()).place(x=330, y=360)

        self.v = tkinter.IntVar(value=0)
        self.canva1.create_rectangle(15, 210, 150, 290, outline="black")  ## rectangle du haut, pour mettre titre
        self.canva1.create_text(75, 230, text="Veuillez choisir l'UE:")
        self.case1 = tkinter.Radiobutton(variable=self.v, value=10)
        self.case2 = tkinter.Radiobutton(variable=self.v, value=20)
        self.case1.config(text="UE5")
        self.case2.config(text="Autre UE")
        self.case1.place(x=20,y=240)
        self.case2.place(x=20,y=260)
        self.canva1.create_text(30, 390, text=self.Qcm.version)

    def AskCorrFile(self):

        FileC = filedialog.askopenfilename(initialdir="Desktop/", title="Feuille de correction")
        print(FileC)
        if "jpg" in FileC or "jpeg" in FileC:
            if FileC != "":
                self.canva1.create_oval(360, 60, 390, 90, fill = 'green')
                return FileC
            else:
                self.canva1.create_oval(360, 60, 390, 90, fill = 'red')
                return ""
        else:
            if ".pdf" in FileC:
                self.AffichageMessage("Ce n'est pas le bon format de fichier. \n Actuellement au format PDF, il faut le format .JPEG ou .JPG")
            else:
                self.AffichageMessage("Ce n'est pas le bon format de fichier\n Il faut un format .JPEG ou .JPG")
            self.canva1.create_oval(360, 60, 390, 90, fill='red')
            return ""


    def AskPackFile(self):
        FileP = filedialog.askdirectory(initialdir="Desktop/", title="Paquet de copie")
        if FileP != "":
            self.canva1.create_oval(360,105,390,135, fill = 'green')  ## voyant check Paquet de copie
            return FileP
        else:
            self.canva1.create_oval(360,105,390,135, fill = 'red')  ## voyant check Paquet de copie
            return ""

    def AskNumber(self):

        nombre = self.Entry1.get()
        FON_Police = tkinter.font.Font(size=20)
        self.canva2.destroy()
        self.canva2 = tk.Canvas(self, width=30, height=30)
        FON_Police = tkinter.font.Font(size=20)
        self.canva2.create_text(15, 15, text=nombre, font=FON_Police)  ## Titre dans le rectangle
        self.canva2.place(x=350, y=170)
        self.Entry1.delete(0, tk.END)
        return nombre

    def AskExcelFile(self):
        if self.Qcm.done==False:
            print("tartine")
            self.AffichageMessage("Veuillez effectuer une correction d'abord")
        else:
            fileE = filedialog.asksaveasfilename(initialdir="Desktop/", title="Où telecharger le classement")
            if fileE != "":
                self.Qcm.toExcel(fileE)
                self.AffichageMessage("Tout a été fait, le classement a été sauvegardé. \nVous pouvez fermer le programme :)")

            else:
                self.AffichageMessage("Telechargement excel annulé")



    def AffichageMessage(self,Message):
        newWindow = tk.Toplevel(app)
        newWindow.geometry('300x100')
        newWindow.resizable(width=0, height=0)
        newWindow.iconbitmap("Logo valid.ico")
        newWindow.title("Qcm correction automatique Probleme")
        label = tk.Label(newWindow, text=Message,).place(x=20,y=20)
        boutonok = tk.Button(newWindow, text="OK", command =lambda :newWindow.destroy()).place(x=135,y=65)



    def TestCorrection(self):
        if (self.Qcm.corrfile==""):
            self.AffichageMessage("Vous n'avez pas donner de Correction")

        elif(self.Qcm.paquetfile==""):
            self.AffichageMessage("Vous n'avez pas donner de Paquet de Copie")

        elif (self.Qcm.nbquestion==0 or self.Qcm.nbquestion== "0"):
            self.AffichageMessage("Le nombre de question est fixé à 0")
        elif(self.Qcm.corrfile!="" and self.Qcm.paquetfile != "" and self.Qcm.nbquestion!=0):
            if self.Qcm.tabresultat[0][0] != 0:
                self.AffichageMessage("Correction déjà effectuée, \n Si vous voulez effectuer une nouvelle corerction, relancez le programme")
            else:
                if self.v.get()==0:
                    self.AffichageMessage("Veuillez choisir une UE")
                elif self.v.get() == 10 :
                   self.Qcm.ChangerUE(True)
                   self.Qcm.Correction()
                   self.AffichageMessage("Correction effectuée")
                   ## print(" True has done") debug
                elif self.v.get()==20:
                    self.Qcm.ChangerUE(False)
                    self.Qcm.Correction()
                    self.AffichageMessage("Correction effectuée")
                    ## print(" False has done") debug


if __name__ == "__main__":
    app = Application()
    app.mainloop()